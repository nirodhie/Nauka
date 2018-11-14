import win32com.client
from openpyxl import Workbook
plikExcela = Workbook()
subject = "Onboarding (Wroclaw) - 01.11.2018"
skoroszyt = plikExcela.create_sheet(subject)

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

#inbox = outlook.GetDefaultFolder(6) # "6" refers to the index of a folder - in this case,
                                    # the inbox. You can change that number to reference
                                    # any other folder

inbox = outlook.Folders['Lukasz.Teska@bd.com'].Folders['Skrzynka odbiorcza'] #.Folders['New users']
messages = inbox.Items
message = messages.GetLast()

subject = "Onboarding (Wroclaw) - 01.11.2018"
skoroszyt = plikExcela.create_sheet(subject)
cellIteration = 1



while message:
    if subject in message.Subject and "RE:" not in message.Subject and "FW:" not in message.Subject:
        realName = message.subject.lstrip(subject)
           # usun czesc tytulu ktora poszukujemy bo jest dla wszystkich taka sama
        position = (message.body.splitlines()[20])
        manager = (message.body.splitlines()[24])
        globalID = (message.body.splitlines()[28])
        costCenter = (message.body.splitlines()[32])

        skoroszyt.cell(cellIteration, 1, realName)
        skoroszyt.cell(cellIteration, 2, position)
        skoroszyt.cell(cellIteration, 3, manager)
        skoroszyt.cell(cellIteration, 4, globalID)
        skoroszyt.cell(cellIteration, 5, costCenter)
        print(realName, position, manager, globalID, costCenter)
        cellIteration = cellIteration + 1
    message = messages.GetPrevious()
plikExcela.save(r'c:\Onboarding_users.xlsx')

